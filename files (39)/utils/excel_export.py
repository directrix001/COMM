"""
utils/excel_export.py
─────────────────────
Multi-sheet Excel export builder used by Tab 2.
"""

from __future__ import annotations

import io
from typing import List, Optional

import numpy as np
import pandas as pd

from utils.formatting import fmt_num_full, fmt_pct, variance_color_class


def build_excel_export(
    leaf_df: pd.DataFrame,
    group_fields: List[str],
    header_a: str,
    header_b: str,
    sel_period: str,
    sel_markets: List[str],
    sel_regions: List[str],
    sel_divisions: List[str],
    scenario_a: str,
    scenario_b: str,
    favorable_is_lower: bool,
    pivot_source_long: Optional[pd.DataFrame],
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── README sheet ──────────────────────────────────────────────────────────
    ws_readme = wb.active
    ws_readme.title = "README"
    readme_data = [
        ("Period", sel_period),
        ("Markets", ", ".join(sel_markets) if sel_markets else "All"),
        ("Regions", ", ".join(sel_regions) if sel_regions else "All"),
        ("Scenario A", header_a),
        ("Scenario B", header_b),
        ("Favorable when", "A < B (costs)" if favorable_is_lower else "A > B (revenue)"),
        ("Division_Desc filter", ", ".join(sel_divisions) if sel_divisions else "All"),
    ]
    for r_idx, (k, v) in enumerate(readme_data, 1):
        ws_readme.cell(r_idx, 1, k).font = Font(bold=True)
        ws_readme.cell(r_idx, 2, v)
    ws_readme.column_dimensions["A"].width = 28
    ws_readme.column_dimensions["B"].width = 50

    # ── Variance (Flat) sheet ─────────────────────────────────────────────────
    ws_flat = wb.create_sheet("Variance (Flat)")
    flat = leaf_df.rename(columns={"A": header_a, "B": header_b})
    flat_cols = group_fields + [header_a, header_b, "Δ (A-B)", "Δ% vs B"]
    flat_out = flat[[c for c in flat_cols if c in flat.columns]]

    hdr_fill = PatternFill(start_color="0F1F3D", end_color="0F1F3D", fill_type="solid")
    hdr_font = Font(color="E8EDF8", bold=True)

    for ci, col in enumerate(flat_out.columns, 1):
        c = ws_flat.cell(1, ci, col)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center")

    red_fill   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    for ri, row in enumerate(flat_out.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            ws_flat.cell(ri, ci, val)

    nrows_flat = ws_flat.max_row
    d_col = flat_out.columns.tolist().index("Δ (A-B)") + 1 if "Δ (A-B)" in flat_out.columns else None
    p_col = flat_out.columns.tolist().index("Δ% vs B") + 1 if "Δ% vs B" in flat_out.columns else None

    def add_cf_rules(ws, cidx, nrows, fav_lower):
        cl = get_column_letter(cidx)
        rng = f"{cl}2:{cl}{nrows}"
        if fav_lower:
            ws.conditional_formatting.add(rng, CellIsRule("greaterThan", ["0"], fill=red_fill))
            ws.conditional_formatting.add(rng, CellIsRule("lessThan",    ["0"], fill=green_fill))
        else:
            ws.conditional_formatting.add(rng, CellIsRule("greaterThan", ["0"], fill=green_fill))
            ws.conditional_formatting.add(rng, CellIsRule("lessThan",    ["0"], fill=red_fill))

    if d_col:
        add_cf_rules(ws_flat, d_col, nrows_flat, favorable_is_lower)
    if p_col:
        add_cf_rules(ws_flat, p_col, nrows_flat, favorable_is_lower)

    for col in ws_flat.columns:
        ws_flat.column_dimensions[get_column_letter(col[0].column)].width = 18
    ws_flat.freeze_panes = "A2"

    # ── Pivot (Outline) sheet ─────────────────────────────────────────────────
    ws_piv = wb.create_sheet("Pivot (Outline)")

    def sums(part):
        a = float(part["A"].sum())
        b = float(part["B"].sum())
        d = a - b
        p = (d / b) if b != 0 else float("nan")
        return a, b, d, p

    df_sorted = leaf_df.sort_values(group_fields).reset_index(drop=True)
    piv_cols = group_fields + [header_a, header_b, "Δ (A-B)", "Δ% vs B"]
    for ci, col in enumerate(piv_cols, 1):
        c = ws_piv.cell(1, ci, col)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center")

    row_cursor = [2]
    spans = []

    GRAND_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    GRAND_FONT = Font(color="E8EDF8", bold=True)
    LVL1_FILL  = PatternFill(start_color="DDE8F8", end_color="DDE8F8", fill_type="solid")
    LVL2_FILL  = PatternFill(start_color="EEF4FC", end_color="EEF4FC", fill_type="solid")
    SUB_FILL   = PatternFill(start_color="F0F5FD", end_color="F0F5FD", fill_type="solid")
    BOLD       = Font(bold=True)
    ITALIC     = Font(italic=True)
    n_dim = len(group_fields)

    def write_row(values, fill=None, font=None, indent_col=None, indent_val=0):
        ri = row_cursor[0]
        for ci, val in enumerate(values, 1):
            c = ws_piv.cell(ri, ci, val)
            if fill:
                c.fill = fill
            if font:
                c.font = font
            if indent_col and ci == indent_col:
                c.alignment = Alignment(indent=indent_val)
            if ci > n_dim:
                c.number_format = "#,##0.00" if ci < n_dim + 5 else "0.00%"
                c.alignment = Alignment(horizontal="right")
        row_cursor[0] += 1
        return ri

    A_t, B_t, D_t, P_t = sums(df_sorted)
    grand_vals = ["GRAND TOTAL"] + [""] * (n_dim - 1) + [A_t, B_t, D_t, P_t]
    write_row(grand_vals, fill=GRAND_FILL, font=GRAND_FONT)

    if n_dim == 1:
        col = group_fields[0]
        for val in df_sorted[col].astype(str).fillna("").unique():
            part = df_sorted[df_sorted[col].astype(str).fillna("") == val]
            a, b, d, p = sums(part)
            write_row([val, a, b, d, p])
    else:
        def recurse_xl(sub_df, level):
            col = group_fields[level]
            is_second_last = level == n_dim - 2
            fill = LVL1_FILL if level == 0 else LVL2_FILL

            for val in sub_df[col].astype(str).fillna("").unique():
                part = sub_df[sub_df[col].astype(str).fillna("") == val]
                hdr_vals = [""] * n_dim + [None, None, None, None]
                hdr_vals[level] = val
                write_row(hdr_vals, fill=fill, font=BOLD, indent_col=level + 1, indent_val=level)

                child_start = row_cursor[0]

                if is_second_last:
                    last_col = group_fields[-1]
                    for lv in part[last_col].astype(str).fillna("").unique():
                        part2 = part[part[last_col].astype(str).fillna("") == lv]
                        a, b, d, p = sums(part2)
                        det_vals = [""] * n_dim + [a, b, d, p]
                        det_vals[n_dim - 1] = lv
                        write_row(det_vals, indent_col=n_dim, indent_val=level + 1)
                else:
                    recurse_xl(part, level + 1)

                child_end = row_cursor[0] - 1
                if child_end >= child_start:
                    spans.append((child_start, child_end, level + 1))

                a, b, d, p = sums(part)
                sub_vals = [""] * n_dim + [a, b, d, p]
                sub_vals[level] = f"{val} — Total"
                write_row(sub_vals, fill=SUB_FILL, font=ITALIC, indent_col=level + 1, indent_val=level)

        recurse_xl(df_sorted, 0)

    nrows_piv = ws_piv.max_row
    d_piv = n_dim + 3
    p_piv = n_dim + 4
    if nrows_piv > 1:
        add_cf_rules(ws_piv, d_piv, nrows_piv, favorable_is_lower)
        add_cf_rules(ws_piv, p_piv, nrows_piv, favorable_is_lower)

    ws_piv.sheet_properties.outlinePr.summaryBelow = True
    for start_r, end_r, outline_level in spans:
        ws_piv.row_dimensions.group(start_r, end_r, outline_level=outline_level, hidden=True)

    ws_piv.freeze_panes = "A2"
    for col in ws_piv.columns:
        ws_piv.column_dimensions[get_column_letter(col[0].column)].width = 20

    # ── PivotSource sheet ─────────────────────────────────────────────────────
    if pivot_source_long is not None and not pivot_source_long.empty:
        ws_src = wb.create_sheet("PivotSource")
        src_cols = list(pivot_source_long.columns)
        for ci, col in enumerate(src_cols, 1):
            c = ws_src.cell(1, ci, col)
            c.fill = hdr_fill
            c.font = hdr_font
        for ri, row in enumerate(pivot_source_long.itertuples(index=False), 2):
            for ci, val in enumerate(row, 1):
                ws_src.cell(ri, ci, val)
        for col in ws_src.columns:
            ws_src.column_dimensions[get_column_letter(col[0].column)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
