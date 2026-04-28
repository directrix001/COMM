"""
services/excel_export.py
─────────────────────────
Multi-sheet Excel export — pure Python, no Streamlit dependency.
Ported from utils/excel_export.py.
"""

from __future__ import annotations

import io
from typing import List, Optional
import numpy as np
import pandas as pd


def build_excel_export(
    leaf_df: pd.DataFrame,
    group_fields: List[str],
    header_a: str,
    header_b: str,
    sel_period: str,
    sel_markets: List[str],
    sel_regions: List[str],
    sel_divisions: List[str],
    scenario_a: str,
    scenario_b: str,
    favorable_is_lower: bool,
    pivot_source_long: Optional[pd.DataFrame],
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # README
    ws_readme = wb.active
    ws_readme.title = "README"
    readme_data = [
        ("Period", sel_period),
        ("Markets", ", ".join(sel_markets) if sel_markets else "All"),
        ("Regions", ", ".join(sel_regions) if sel_regions else "All"),
        ("Scenario A", header_a),
        ("Scenario B", header_b),
        ("Favorable when", "A < B (costs)" if favorable_is_lower else "A > B (revenue)"),
        ("Division_Desc filter", ", ".join(sel_divisions) if sel_divisions else "All"),
    ]
    for r_idx, (k, v) in enumerate(readme_data, 1):
        ws_readme.cell(r_idx, 1, k).font = Font(bold=True)
        ws_readme.cell(r_idx, 2, str(v))
    ws_readme.column_dimensions["A"].width = 28
    ws_readme.column_dimensions["B"].width = 50

    hdr_fill  = PatternFill(start_color="0F1F3D", end_color="0F1F3D", fill_type="solid")
    hdr_font  = Font(color="E8EDF8", bold=True)
    red_fill  = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    grn_fill  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    def add_cf(ws, cidx, nrows, fav_lower):
        cl  = get_column_letter(cidx)
        rng = f"{cl}2:{cl}{nrows}"
        if fav_lower:
            ws.conditional_formatting.add(rng, CellIsRule("greaterThan", ["0"], fill=red_fill))
            ws.conditional_formatting.add(rng, CellIsRule("lessThan",    ["0"], fill=grn_fill))
        else:
            ws.conditional_formatting.add(rng, CellIsRule("greaterThan", ["0"], fill=grn_fill))
            ws.conditional_formatting.add(rng, CellIsRule("lessThan",    ["0"], fill=red_fill))

    # Variance (Flat)
    ws_flat  = wb.create_sheet("Variance (Flat)")
    flat     = leaf_df.rename(columns={"A": header_a, "B": header_b, "delta": "Δ (A-B)", "delta_pct": "Δ% vs B"})
    flat_cols = group_fields + [header_a, header_b, "Δ (A-B)", "Δ% vs B"]
    flat_out  = flat[[c for c in flat_cols if c in flat.columns]]

    for ci, col in enumerate(flat_out.columns, 1):
        c = ws_flat.cell(1, ci, col)
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = Alignment(horizontal="center")

    for ri, row in enumerate(flat_out.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            ws_flat.cell(ri, ci, val)

    nrows_flat = ws_flat.max_row
    cols_list  = flat_out.columns.tolist()
    d_col = cols_list.index("Δ (A-B)") + 1 if "Δ (A-B)" in cols_list else None
    p_col = cols_list.index("Δ% vs B") + 1 if "Δ% vs B" in cols_list else None
    if d_col: add_cf(ws_flat, d_col, nrows_flat, favorable_is_lower)
    if p_col: add_cf(ws_flat, p_col, nrows_flat, favorable_is_lower)
    for col in ws_flat.columns:
        ws_flat.column_dimensions[get_column_letter(col[0].column)].width = 18
    ws_flat.freeze_panes = "A2"

    # PivotSource
    if pivot_source_long is not None and not pivot_source_long.empty:
        ws_src = wb.create_sheet("PivotSource")
        for ci, col in enumerate(pivot_source_long.columns, 1):
            c = ws_src.cell(1, ci, col); c.fill = hdr_fill; c.font = hdr_font
        for ri, row in enumerate(pivot_source_long.itertuples(index=False), 2):
            for ci, val in enumerate(row, 1):
                ws_src.cell(ri, ci, val)
        for col in ws_src.columns:
            ws_src.column_dimensions[get_column_letter(col[0].column)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
